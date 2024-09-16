import pandas as pd
import PIL.ImageFont as ImageFont
import qrcode
import os
import shutil
from nicegui import ui
from typing import Union
from PIL import Image, ImageDraw
from fpdf import FPDF
from openpyxl import load_workbook
import multiprocessing

lock = multiprocessing.Lock()

def create_directories():
    if not os.path.exists('barcodes'):
        os.makedirs('barcodes')
    if not os.path.exists('download_temp'):
        os.makedirs('download_temp')


def debug_print(message):
    """
    Prints a debug message in green color.

    :param message: The debug message.
    :type message: str
    """
    print("\033[92m" + str(message) + "\033[0m")

class excelWriter:
    file_lock = multiprocessing.Lock()

    def __init__(self):
        self.wb = None
        self.sheet = None
        self.i = 1
        self.done = False
        self.file_path = "Verbrauchsmaterial_ELab_TRGE.xlsx"
        self.sheet_name = "Verbrauchsmaterial"

    def read_excel(self, lock):
        with lock:
            debug_print("Locked File")
            try:
                return pd.read_excel('Verbrauchsmaterial_ELab_TRGE.xlsx', sheet_name='Verbrauchsmaterial', dtype={'Available':bool}, usecols={"id","Available","Name"})
            except Exception as e:
                debug_print("Error Reading File! Error:\n"+str(e))
                ui.notify('Error Loading File!', category='error')
                return None

    def update_excel(self, input: Union[str, list[dict[str, int]]], setting:bool=False, lock=None):
        with lock:
            debug_print("Locked File")
            try:
                self.wb = load_workbook(self.file_path)
                self.sheet = self.wb[self.sheet_name]
                if type(input) == str:
                    debug_print("Updating String")

                    for self.row in self.sheet.iter_rows(values_only=True):
                        if self.row[0] != None:
                            if self.row[0] == int(input):
                                self.sheet['B'+str(self.i)] = setting
                                debug_print("Edited Value from ID")
                                break
                        self.i += 1
                    
                    self.wb.save(self.file_path)
                elif type(input) == list:
                    debug_print("Updating List")
                    for self.inp in input:
                        self.i = 1

                        for row in self.sheet.iter_rows(values_only=True):
                            if row[0] != None:
                                if row[0] == int(self.inp['id']):
                                    self.sheet['B'+str(self.i)] = setting
                                    debug_print("Edited Value from ID")
                                    break
                            self.i += 1

                    self.wb.save(self.file_path)
                self.wb.close()
                debug_print("Written into Excel")
            except Exception as e:
                ui.notify('Error Updating Availability!', category='error')
                debug_print("Error Updating Availability! Error:\n"+str(e))
                self.wb.save(self.file_path)
                self.wb.close()

        self.wb = None
        debug_print("Updated Excel")

class InventoryManager:

    """
    A class for managing an inventory of items with unique serial numbers, names, descriptions, 
    and availability statuses.
    """

    qr_width = 1000

    def __init__(self):
        """
        Initializes the InventoryManager with the inventory data from the Excel file 
        Bauteileschrank.xlsx.
        """
        create_directories()
        self.running_data = None
        self.table = None
        self.input = None
        self.wb = None
        self.sht = None

    """def save_data(self):
        """""""
        Saves the current inventory data to existing Excel file.
        Runningdata contains the columns:
        id: The serial number of the item.
        Available: The availability status of the item.
        Name: The name of the item.
        """"""
        try:
            dataframe = pd.read_excel('Verbrauchsmaterial_ELab_TRGE.xlsx', sheet_name='Verbrauchsmaterial', dtype={'Available':bool})
            debug_print(dataframe)
        except Exception as e:
            print("Error Saving File! Error:\n" + str(e))
            ui.notify('Error Saving File!', category='error')"""

    def read_data(self):
        """
        Reads the inventory data from an Excel file.

        :return: The inventory data.
        :rtype: pd.DataFrame
        """
        xlWriter = excelWriter()
        ret_data = xlWriter.read_excel(lock)
        if not ret_data.empty:
            self.running_data = ret_data
        

    """def gen_new_serial(self) -> int:
        """"""
        Generates a new serial number for the inventory.

        :return: The new serial number.
        :rtype: int
        """"""
        lastID = self.running_data['id'].max()
        debug_print("Generated new Serial: "+str(lastID + 1))
        return lastID + 1"""

    def update_availability(self, input: Union[str, list[dict[str, int]]], setting:bool=False):
        """
        Updates the availability of an item or a list of items.

        :param input: The serial number or list of items.
        :type input: Union[str, List[Dict[str, int]]
        :param setting: The availability setting.
        :type setting: bool
        """
        input_copy = input[:]
        debug_print(input_copy)
        xlWriter = excelWriter()
        proc = multiprocessing.Process(target=xlWriter.update_excel, args=(input_copy, setting, lock))
        proc.start()
        if input != None:
            if type(input) == str:
                debug_print("Updating String")
                self.running_data.loc[self.running_data['id']==int(input_copy), 'Available'] = setting
                self.table.update_rows(self.running_data.loc[:].to_dict('records'))
            elif type(input) == list:
                debug_print("Updating List")
                for inp in input_copy:
                    self.running_data.loc[self.running_data['id']==inp['id'], 'Available'] = setting
                    self.table.update_rows(self.running_data.loc[:].to_dict('records'))
            debug_print("Updated Table")
        
    def update_data(self):
        """
        Updates the data in the table.
        """
        self.read_data()
        self.table.update_rows(self.running_data.loc[:].to_dict('records'))
        self.table.update()
        debug_print("Updated Data")

    def gen_qr_code(self, serial:int, name:str, sap:int=0):
        filename = f"barcodes/{serial}.pdf"

        debug_print("Generating QR-Code for ID: "+str(serial))
        debug_print("Name: "+name)

        # Generate the QR code
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(id)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        # Create a new image with the QR code and text
        width, height = img.size
        new_width = width + 1071
        new_height = height + 50
        new_img = Image.new("RGB", (new_width, new_height), "white")
        new_img.paste(img, (0, 0))

        # Add text to the image
        draw = ImageDraw.Draw(new_img)
        font = ImageFont.truetype("arial.ttf", 100)  # You may need to install a font
        if len(name) < 33:
            draw.text((width + 10, 40), name, font=font, fill="black")
            draw.text((width + 10, 180), sap, font=font, fill="black")
        elif len(name) < 66:
            draw.text((width + 10, 40), name[:33], font=font, fill="black")
            draw.text((width + 10, 120), name[33:], font=font, fill="black")
            draw.text((width + 10, 180), sap, font=font, fill="black")
        elif len(name) < 99:
            draw.text((width + 10, 40), name[:33], font=font, fill="black")
            draw.text((width + 10, 120), name[33:66], font=font, fill="black")
            #draw.text((width + 10, 180), name[99:], font=font, fill="black")
            draw.text((width + 10, 180), sap, font=font, fill="black")
        #elif len(name) < 132:
            #draw.text((width + 10, 40), name[:33], font=font, fill="black")
            #draw.text((width + 10, 120), name[33:66], font=font, fill="black")
            #draw.text((width + 10, 180), sap, font=font, fill="black")
            #draw.text((width + 10, 230), name[99:132], font=font, fill="black")
        else:
            ui.notify("Name is too long lengt:" + str(len(name)))

        draw.text((50, 300), str(serial), font=font, fill="black")

        #new_img.save(f"barcodes/{serial}_small.png")

        debug_print(new_img.size)

        middle_img = Image.new("RGB", (2156, 593), "white")
        middle_img.paste(new_img, (367, 38))
        #middle_img.save(f"barcodes/{serial}_mid.png")
                        
        big_img = Image.new("RGB", (2156, 728), "white")
        big_img.paste(new_img, (612, 38))
        #big_img.save(f"barcodes/{serial}_big.png")

        pdf = FPDF()
        pdf.set_line_width(0.5)
        pdf.add_page()
        pdf.rect(10, 10, 53, 14.5)
        pdf.image(new_img, x=0, y=10, w=53, keep_aspect_ratio=True)
        pdf.rect(10, 70, 82, 22)
        pdf.image(middle_img, x=0, y=70, w=82, keep_aspect_ratio=True)
        pdf.rect(10, 150, 82, 27)
        pdf.image(big_img, x=0, y=150, w=82, keep_aspect_ratio=True)
        pdf.output(f"barcodes/{serial}.pdf")

    def download_selected_qr_codes(self, id_list):
        """
        Downloads a QR code for each item with the given IDs.

        :param ids: The IDs of the items.
        :type ids: List[int]
        """
        for serial in id_list:
            shutil.copy(f"barcodes/{serial['id']}.pdf", f"download_temp/{serial['id']}.pdf")
        ui.download(shutil.make_archive('labels', 'zip', 'download_temp'))
        for filename in os.listdir('download_temp'):
            os.remove(f"download_temp/{filename}")
    
    def save_uploaded_file(self, file):
        """
        Saves an uploaded file to the inventory.

        :param file: The uploaded file.
        :type file: File
        """
        try:
            with open(file.name, 'wb') as f:
                f.write(file.content.read())
            debug_print("Uploaded File")
            self.update_data()
            for id in self.running_data['id']:
                if not os.path.exists(f"barcodes/{id}.png"):
                    self.gen_qr_code(id, self.running_data.loc[self.running_data['id']==id, 'Name'].values[0])
        except Exception as e:
            ui.notify('Error Uploading File!', category='error')
            debug_print(f"Error Uploading File! Error:{e}")

    def up_download_excel(self):
        """
        Uploads or downloads an Excel file with the inventory data. When Downloading it also exports the QR-Codes.
        """
        with ui.dialog() as dialog, ui.card():
            with ui.tabs() as tabs:
                ui.tab('up', label='Upload', icon='upload')
                ui.tab('down', label='Download', icon='download')
            with ui.tab_panels(tabs, value='up'):
                with ui.tab_panel('up'):
                    ui.upload(multiple=False, auto_upload=True, on_upload= lambda upload: self.save_uploaded_file(upload)).props('accept=.xlsx')
                with ui.tab_panel('down'):
                    ui.button('Download Excel', on_click=lambda: ui.download('Verbrauchsmaterial_ELab_TRGE.xlsx'))
                    #ui.button('Download QR-Codes', on_click=lambda: ui.download(shutil.make_archive('barcodes', 'zip', 'barcodes')))


        dialog.open()

    def run(self):
        """
        Starts the inventory management application.

        The application will run until it is manually stopped.
        """
        ui.run(port=80,title='CoTrack',dark=None)

inv = InventoryManager()
inv.read_data()

@ui.page('/')
def normal_view():
    """
    Displays the normal view of the inventory.
    """
    inv.table =  ui.table.from_pandas(inv.running_data).classes('w-full')
    inv.table.columns[1]['sortable'] = True
    inv.table.columns[0]['sortable'] = True
    with inv.table.add_slot('top-left'):
        inv.inputRef = ui.input(placeholder='Scanner').on('keydown.enter',lambda: (inv.update_availability(inv.inputRef.value, False), inv.inputRef.set_value(None))).props('type=search')
        #ui.button('Clear').on_click(lambda: (inv.inputRef.set_value(None)))
    with inv.table.add_slot('top-right'):
        with ui.link(target=editor_view):
            ui.button('Editor View')
    inv.table.set_fullscreen(True)
    inv.table.add_slot('body-cell-Available', '''
        <q-td key="Available" :props="props">
            <q-badge :color="props.value < true ? 'red' : 'green'">
                {{ props.value < true ? 'No' : 'Yes' }}
            </q-badge>
        </q-td>
        ''')
    ui.run_javascript('document.querySelector("input").focus()')

@ui.page('/editor')
def editor_view():
        """
        Displays the editor view of the inventory.
        """
        inv.table = ui.table.from_pandas(inv.running_data, selection='multiple').classes('w-full')
        inv.table.columns[1]['sortable'] = True
        inv.table.columns[0]['sortable'] = True
        with inv.table.add_slot('top-left'):
            inputRef = ui.input(placeholder='Search').on('keydown.enter',lambda: (inv.update_availability(inputRef.value, True), inputRef.set_value(None))).props('type=search')
            inv.table.bind_filter_from(inputRef, 'value')
            with inputRef.add_slot("append"):
                ui.icon('search')
        with inv.table.add_slot('top-right'):
            ui.button('Refresh',on_click=lambda: inv.update_data())
            ui.button('Upload/Download Excel', on_click=lambda: inv.up_download_excel())
            ui.button('Download Label(s)', on_click=lambda: inv.download_selected_qr_codes(inv.table.selected)).bind_enabled_from(inv.table, 'selected', backward=lambda val: bool(val))
            ui.button('Refill', on_click=lambda: (inv.update_availability(inv.table.selected, True), ui.update(inv.table))).bind_enabled_from(inv.table, 'selected', backward=lambda val: bool(val))
            #ui.button('Remove', on_click=lambda: (inv.delete_part(inv.table.selected))).bind_enabled_from(inv.table, 'selected', backward=lambda val: bool(val))
            with ui.link(target=normal_view):
                ui.button('Close View')

        #with inv.table.add_slot('bottom'):
        #    with inv.table.row():
        #        with inv.table.cell():
        #            ui.button(on_click=lambda: (
        #                inv.add_part(new_name.value),
        #                new_name.set_value(None),
        #            ), icon='add').props('flat fab-mini')
        #        with inv.table.cell():
        #            new_name = ui.input('Name')
        inv.table.set_fullscreen(True)
        inv.table.add_slot('body-cell-Available', '''
            <q-td key="Available" :props="props">
                <q-badge :color="props.value < true ? 'red' : 'green'">
                    {{ props.value < true ? 'No' : 'Yes' }}
                </q-badge>
            </q-td>
            ''')

inv.run()